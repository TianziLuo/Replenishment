import subprocess
from pathlib import Path
from openpyxl import load_workbook

def find_folder(keyword: str) -> str:
    go_exe = r"C:\Users\monica\Desktop\Seagull\Replenishment\findfolder.exe"
    proc = subprocess.run(
        ["search_folder.exe", keyword],
        text=True,
        capture_output=True
    )
    if proc.returncode != 0:
        raise RuntimeError(f"Go error: {proc.stderr}")
    return proc.stdout.strip()

def readSKU(folder: str) -> str:
    sku_file = Path(folder) / "sku.txt"
    if sku_file.exists():
        with open(sku_file, "r", encoding="utf-8") as f:
            sku_value = f.readline().strip()
        print("SKU:", sku_value)
        return sku_value
    else:
        print("⚠️ sku.txt not found")
        return None

def writeSKUtoExcel(sku: str):
    excel_path = r"C:\Frank\补出库单.xlsx"
    wb = load_workbook(excel_path)
    sheet = wb["登记"]

    # 找 D 列第一个空白单元格
    col = 4  # D列
    row = 1
    while sheet.cell(row=row, column=col).value is not None:
        row += 1

    # 写入 SKU
    sheet.cell(row=row, column=col, value=sku)
    wb.save(excel_path)
    print(f"SKU {sku} 已写入 {excel_path} 的 D{row}")

if __name__ == "__main__":
    keyword = input("Please scan the keyword: ")
    folder = find_folder(keyword)
    if folder:
        print("Found folder:", folder)
        sku = readSKU(folder)
        if sku:
            writeSKUtoExcel(sku)
    else:
        print("No matching folder found")
