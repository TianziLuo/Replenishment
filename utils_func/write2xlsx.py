import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from readTXT import readSKU, readOrderNO
import os


def get_last_row(ws):
    """
    Find the last row with actual content.
    """
    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row=row, column=c).value is not None for c in range(1, ws.max_column + 1)):
            return row
    return 0


def writeSKU2Excel(folder: Path, sticker: str):
    sku_value = readSKU(folder)
    order_no = readOrderNO(folder)

    if not sku_value and not order_no:
        print("⚠️ No data read, operation aborted")
        return

    excel_path = Path(r"C:\Frank\补发货.xlsx")
    wb = load_workbook(excel_path)
    ws = wb["InHand"]

    last_row = get_last_row(ws)
    font_style = Font(name="DengXian", size=11)

    ws.cell(row=last_row + 1, column=1, value=sku_value).font = font_style
    ws.cell(row=last_row + 1, column=2, value=order_no).font = font_style
    ws.cell(row=last_row + 1, column=3, value=sticker).font = font_style

    wb.save(excel_path)
    print(f"✅ Write completed: SKU={sku_value}, OrderNO={order_no}, Sticker={sticker}")

    os.startfile(excel_path)

if __name__ == "__main__":
    # Expect sticker as sys.argv[2] from Go script
    if len(sys.argv) >= 3:
        sticker = sys.argv[2]
    else:
        sticker = ""

    writeSKU2Excel(sticker)

'''
if __name__ == "__main__":
    writeSKU2Excel()
'''
