import subprocess
from pathlib import Path
from openpyxl import load_workbook

def find_folder(keyword: str) -> str:
    proc = subprocess.run(
        ["search_folder.exe", keyword],
        text=True,
        capture_output=True
    )
    if proc.returncode != 0:
        raise RuntimeError(f"Go error: {proc.stderr}")
    return proc.stdout.strip()

def readOrder(folder: str) -> str:
    order_file = Path(folder) / "orderno.txt"
    if order_file.exists():
        with open(order_file, "r", encoding="utf-8") as f:
            orderNo_value = f.readline().strip()
        print("OrderNo:", orderNo_value)
        return orderNo_value
    else:
        print("⚠️ orderno.txt not found")
        return None

def writeOrder2Excel(orderNo: str):
    excel_path = r"C:\Frank\补发货.xlsx"
    excel_file = Path(excel_path)
    if not excel_file.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_file}")

    wb = load_workbook(excel_path)
    sheet = wb["InHand"]

    # 找 B 列第一个空白单元格（从第2行开始）
    col = 2  # B列
    row = 2
    while sheet.cell(row=row, column=col).value is not None:
        row += 1

    # 写入 orderNo
    sheet.cell(row=row, column=col, value=orderNo)
    wb.save(excel_path)
    print(f"orderNo {orderNo} 已写入 {excel_path} 的 B{row}")

if __name__ == "__main__":
    keyword = input("Please scan the keyword: ")
    folder = find_folder(keyword)
    if folder:
        print("Found folder:", folder)
        orderNo = readOrder(folder)
        if orderNo:
            writeOrder2Excel(orderNo)
    else:
        print("No matching folder found")
